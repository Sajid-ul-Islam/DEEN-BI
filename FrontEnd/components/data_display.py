import os
from datetime import datetime
from io import BytesIO

import pandas as pd
import streamlit as st





























def _safe_datetime_series(value) -> pd.Series:
    if value is None:
        return pd.Series(dtype="datetime64[ns]")
    converted = pd.to_datetime(value, errors="coerce")
    if isinstance(converted, pd.Series):
        return converted
    if isinstance(converted, pd.Index):
        return pd.Series(converted)
    if pd.isna(converted):
        return pd.Series(dtype="datetime64[ns]")
    return pd.Series([converted])






def file_summary(uploaded_file, df: pd.DataFrame | None, required_columns: list[str]):
    if not uploaded_file:
        st.info("No file uploaded yet.")
        return False

    st.caption(f"File: {uploaded_file.name}")
    if df is None:
        st.warning("Could not read this file.")
        return False

    c1, c2, c3 = st.columns(3)
    c1.metric("Rows", len(df))
    c2.metric("Columns", len(df.columns))
    c3.metric("Required", len(required_columns))

    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        st.error(f"Missing required columns: {', '.join(missing)}")
        return False
    st.success("Required columns check passed.")
    return True






def export_to_excel(df: pd.DataFrame, sheet_name: str = "Analysis Report") -> bytes:
    """High-fidelity Excel export with professional styling (borders, headers, auto-width)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        wb = writer.book
        ws = wb[sheet_name]
        
        # Styles
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Header Styling
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        # Freeze Header
        ws.freeze_panes = 'A2'
        
        # Auto-adjust column widths
        for i, col_name in enumerate(df.columns):
            col_letter = ws.cell(row=1, column=i+1).column_letter
            max_len = max(df[col_name].astype(str).str.len().max(), len(str(col_name))) + 2
            ws.column_dimensions[col_letter].width = min(max_len, 50)
            
            # Row Styling
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=i+1).border = thin_border

    output.seek(0)
    return output.read()


def show_last_updated(path: str):
    if not os.path.exists(path):
        return
    updated = datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M:%S")
    st.caption(f"Last updated: {updated}")